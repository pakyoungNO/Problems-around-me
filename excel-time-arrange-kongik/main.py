from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.append(["출차시간", "차 종류"])
# 1. 시간단위로 임의의 값을 넣어준다.
# 2. 차 종류도 임의의로 적어준다.
departingtime = [
("06:00:23", "기무찌차"),
("07:23:12", "김치차"),
("12:55:10", "정오차"),
("15:23:12", "외국차"),
("07:34:09", "아침차"),
("08:32:55", "더하기차"),
("06:00:11", "처음차"),
("06:22:55", "다시차"),
("15:43:09", "오후후차"),
("16:34:39", "반차")]

for cell in departingtime:
    ws.append(cell)

# 3. 시간단위의 개수를 세는 방법을 찾는다. 반복문은 아니까. 시간단위 개수세기만 성공하면 된다.
# if ws["A2"].value >= "06:00:00" and ws["A2"].value < "07:00:00":
#     print("hello")
# 위와 같이 수 범위를 세위서 대소비교를 해주면 가능하다. ㅇㅇ

checktime6 = 0
checktime7 = 0
checktime8 = 0
checktime9 = 0
checktime10 = 0
checktime11 = 0
checktime12 = 0
checktime13 = 0
checktime14 = 0
checktime15 = 0
checktime16 = 0
checktime17 = 0


for col in ws.iter_cols(min_row=2, max_row=11, min_col=1, max_col=1): # 튜플형식이 아니라서 값을 변경가능
    for cell in col:
        if cell.value >= "06:00:00" and cell.value < "07:00:00":
            checktime6 = checktime6 + 1
        elif cell.value >= "07:00:00" and cell.value < "08:00:00":
            checktime7 = checktime7 + 1
        elif cell.value >= "08:00:00" and cell.value < "09:00:00":
            checktime8 = checktime8 + 1
        elif cell.value >= "09:00:00" and cell.value < "10:00:00":
            checktime9 = checktime9 + 1
        elif cell.value >= "010:00:00" and cell.value < "11:00:00":
            checktime10 = checktime10 + 1
        elif cell.value >= "011:00:00" and cell.value < "12:00:00":
            checktime11 = checktime11 + 1
        elif cell.value >= "012:00:00" and cell.value < "13:00:00":
            checktime12 = checktime12 + 1
        elif cell.value >= "013:00:00" and cell.value < "14:00:00":
            checktime13 = checktime13 + 1
        elif cell.value >= "014:00:00" and cell.value < "15:00:00":
            checktime14 = checktime14 + 1
        elif cell.value >= "015:00:00" and cell.value < "16:00:00":
            checktime15 = checktime15 + 1
        elif cell.value >= "016:00:00" and cell.value < "17:00:00":
            checktime16 = checktime16 + 1
        elif cell.value >= "017:00:00" and cell.value < "018:00:00":
            checktime17 = checktime17 + 1
        else:
            checktime17 = checktime17
# 4. 개수를 센다면 다른 부분에 표 형식으로 저장한다.

check = (checktime6, checktime7, checktime8, checktime9, checktime10, checktime11,
         checktime12, checktime13, checktime14, checktime15, checktime16, checktime17)

ws["D1"] = "기준시간"
ws["E1"] = "차량수"
num = (6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17)
number = 0
for col in ws.iter_cols(min_row=2, max_row=13, min_col=4, max_col=4): # 튜플형식이 아니라서 값을 변경가능
    for cell in col:
        cell.value = "{}:00:00".format(num[number])
        number = number + 1
number = 0
for col in ws.iter_cols(min_row=2, max_row=13, min_col=5, max_col=5): # 튜플형식이 아니라서 값을 변경가능
    for cell in col:
        cell.value = check[number]
        number = number + 1

wb.save("for-excel-kongik.xlsx")